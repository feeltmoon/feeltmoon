import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill

window = tk.Tk()


window.geometry("670x300")
window.title("Hello!")

style = ttk.Style()
style.theme_use("clam")

def rgb_hack(rgb):
    return "#%02x%02x%02x" % rgb 
window.config(bg=rgb_hack((255, 255, 255)))


def open_file():
    file_types = [('excel file','*.xlsx')]
    filename = fd.askopenfilename(title='Open',initialdir='/',filetypes=file_types)
    showinfo(title='Selected File',message=filename)
    txt.set(filename)


def open_file2():
    file_types = [('excel file','*.xlsx')]
    filename = fd.askopenfilename(title='Open',initialdir='/',filetypes=file_types)
    showinfo(title='Selected File',message=filename)
    txt1.set(filename)
    
    
def get_Entry():
    get_data = entry.get()
    return get_data
    
def get_Entry1():
    get_data = entry1.get()
    return get_data
    

def save_file_and_run():
    # Dynamic Path
    savePath = fd.askdirectory()
    showinfo(title='Selected File',message='HELLO')
    getEntryData = get_Entry()
    getEntryData1 = get_Entry1()
    showinfo(title='GetEntryData',message=getEntryData)
    showinfo(title='GetEntryData1',message=getEntryData1)
    
    # study
    df = pd.read_excel(txt_field.get(),sheet_name='Fields')
    # other
    df1 = pd.read_excel(txt_field1.get(),sheet_name='Fields') 
    
    # use xlsxwriter to create a xlsx file
    outputFile = xlsxwriter.Workbook(savePath + r'\Result.xlsx')
    outputFile.close()
    
    writer = pd.ExcelWriter(savePath + r'\Result.xlsx', engine='xlsxwriter')
    
    if getEntryData != '' and getEntryData1 == '':
        #showinfo(title='prompt',message='Forms are not empty')
        entry_lst = []
        #entry_lst.append(getEntryData.split(,))
        entry_lst = getEntryData.split(',')
        #showinfo(title='prompt',message=entry_lst[0])
        for formOID in entry_lst:
            #showinfo(title='prompt',message=formOID)
            df_Form_s = df.loc[df['FormOID']==formOID,:]
            df_Form_o = df1.loc[df1['FormOID']==formOID,:]
            #print(df_Form_s.info())

            #df_Form.to_excel(savePath + '\\' + formOID + '.xlsx', index=False)
            df_Form_s.to_excel(savePath + '\\' + formOID + '_study' + '.xlsx',index=False)    
            df_Form_o.to_excel(savePath + '\\' + formOID + '_toth' + '.xlsx',index=False)    
            showinfo(title='PROMPT',message='Form Combined to xlsx file')
            
            # read xlsx 
            form_s = pd.read_excel(savePath + '\\' + formOID + '_study' + '.xlsx')
            form_o = pd.read_excel(savePath + '\\' + formOID + '_toth' + '.xlsx')
            
            # write to output
            form_s.to_excel(writer, sheet_name=formOID + '_study',index=False)
            form_o.to_excel(writer, sheet_name=formOID + '_oth',index=False)            
        
        # save the output file    
        writer.save()
        
        # begin compare
        #wb = openpyxl.load_workbook(outputFile)
        # loop through each worksheets
        #print(outputFile.filename)
        wb = openpyxl.load_workbook(outputFile.filename)
#        print(len(wb.sheetnames))
#        sheet_number = len(wb.sheetnames)
#        for i in range(sheet_number):
#            print(i)
        for i in range(len(wb.sheetnames)-1):
            #print(i)
            ws = wb.worksheets[i]
            ws_nxt = wb.worksheets[i+1]
            print(ws.title)
            print(ws_nxt.title)
            if ws.title.split('_')[0] == ws_nxt.title.split('_')[0]:
                showinfo(title='prompt',message='ws.title equal')
                for row in ws.iter_rows():
                    for cell in row:
                        if ws_nxt[cell.coordinate].value != cell.value:
                            #showinfo(title='prompt',message='cell.value not equal')
                            cell.fill = PatternFill("solid", start_color="00FFFF99")
                            #cell.fill = PatternFill(start_color='0055CCCC',end_color='0055CCCC',fill_type='solid')
                            
        wb.save(outputFile.filename)

        
    elif getEntryData != '' and getEntryData1 != '':
        showinfo(title='prompt',message='both lst not empty')
        
        entry_lst_frm = []
        entry_lst_fld = []
        
        entry_lst_frm = getEntryData.split(',')
        entry_lst_fld = getEntryData1.split(',')
                
        print(entry_lst_frm)
        print(entry_lst_fld)
        
#        for i in range(len(entry_lst_frm)):
#            print(i, entry_lst_frm[i])
#            row_flt = df['FormOID']==entry_lst_frm[i] & df['FieldOID']==entry_lst_fld[i]
#            df_Form_s = df.loc[row_flt,:]
#            df_Field_s = df.loc[df['FieldOID']==entry_lst_frm[i],:]
                #df_Form_s = df.loc[df['FormOID']==formOID,:]
        
        #df_append = pd.DataFrame()
        
        for formOID in entry_lst_frm:
            print(formOID)
            #row_flt = df['FormOID']==formOID
            df_append_s = pd.DataFrame()
            df_append_o = pd.DataFrame()
            for fieldOID in entry_lst_fld:
                print(fieldOID)      
                
                row_flt_s = (df['FormOID']==formOID) & (df['FieldOID']==fieldOID)
                row_flt_o = (df1['FormOID']==formOID) & (df1['FieldOID']==fieldOID)
                
                df_FieldRow_s = df.loc[row_flt_s,:]
                df_append_s = df_append_s.append(df_FieldRow_s,sort=False)
                
                df_FieldRow_o = df1.loc[row_flt_o,:]
                df_append_o = df_append_o.append(df_FieldRow_o,sort=False)
                
                
            # df_append saved as xlsx file
            df_append_s.to_excel(savePath + '\\' + formOID + '_study' + '.xlsx',index=False)
            df_append_o.to_excel(savePath + '\\' + formOID + '_toth' + '.xlsx',index=False)      
            # empty df_append
            df_append_s.drop(df_append_s.index , inplace=True)
            df_append_o.drop(df_append_o.index , inplace=True)
            
            # read xlsx file
            excl_s = pd.read_excel(savePath + '\\' + formOID + '_study' + '.xlsx')
            excl_o = pd.read_excel(savePath + '\\' + formOID + '_toth' + '.xlsx')        
            # copy xlsx file content and move to Result.xlsx
            excl_s.to_excel(writer, sheet_name=formOID + '_study',index=False)
            excl_o.to_excel(writer, sheet_name=formOID + '_oth',index=False)
            
#            # empty df_append
#            df_append_s.drop(df_append_s.index , inplace=True)
#            df_append_o.drop(df_append_o.index , inplace=True)
        
        
        # save the output file    
        writer.save()
        
        
        wb = openpyxl.load_workbook(outputFile.filename)
        
        for i in range(len(wb.sheetnames)-1):
            #print(i)
            ws = wb.worksheets[i]
            ws_nxt = wb.worksheets[i+1]
            print(ws.title)
            print(ws_nxt.title)
            if ws.title.split('_')[0] == ws_nxt.title.split('_')[0]:
                showinfo(title='prompt',message='ws.title equal')
                for row in ws.iter_rows():
                    for cell in row:
                        if ws_nxt[cell.coordinate].value != cell.value:
                            #showinfo(title='prompt',message='cell.value not equal')
                            cell.fill = PatternFill("solid", start_color="00FFFF99")
                            ws_nxt[cell.coordinate].fill = PatternFill("solid", start_color="00FFFF99")
                            #cell.fill = PatternFill(start_color='0055CCCC',end_color='0055CCCC',fill_type='solid')                         
        wb.save(outputFile.filename)
        
        

    showinfo(title='PROMPT',message='Done')
    
    
    
    
    
#label_01
myfont = ('Arial', 10)
label = tk.Label(window,text="ALS_First:",width=13,height=2,font=myfont,bg='White')
label.grid(row=1,column=1)
#text_field: to show path    
txt = tk.StringVar(None)    
txt_field = tk.Entry(window, textvariable=txt,bg='Silver')
txt_field.grid(row=1,columns=1)
txt_field.place(x=95, y=33, width=525, height=25)
#button: openFile
browse_text = tk.StringVar()
button = tk.Button(window, text='Open',command=open_file,width=8,height=1,bg="cornflower blue", fg="White")
button.grid(row=1,column=1)
button.place(x=23, y=31)

#label_02
myfont1 = ('Arial', 10)
label1 = tk.Label(window,text="ALS_Second:",width=14,height=1,font=myfont1,bg='White')
label1.grid(row=1,column=1)
label1.place(x=3,y=61)
#text_field: to show path    
txt1 = tk.StringVar(None)    
txt_field1 = tk.Entry(window, textvariable=txt1,bg='Light Gray')
txt_field1.grid(row=1,columns=1)
txt_field1.place(x=95, y=87, width=525, height=25)
#button: openFile
browse_text1 = tk.StringVar()
button1 = tk.Button(window, text='Open',command=open_file2,width=8,height=1,bg="cornflower blue", fg="White")
button1.grid(row=1,column=1)
button1.place(x=23, y=84)


#entry_01
entry = tk.Entry(window,bg='Silver')
entry.grid(row=1,columns=1)
entry.place(x=25, y=137, width=595, height=25)
#entry_01_label
entry_01_label_myfont = ('Arial', 10)
entry_01_label = tk.Label(window,text="Form(s):",width=11,height=1,font=entry_01_label_myfont,bg='White')
entry_01_label.grid(row=1,column=1)
entry_01_label.place(x=3,y=115)


#entry_02
entry1 = tk.Entry(window,bg='Silver')
entry1.grid(row=1,columns=1)
entry1.place(x=25, y=187, width=595, height=25)
#entry_01_label
entry_02_label_myfont = ('Arial', 10)
entry_02_label = tk.Label(window,text="Fields(s):",width=11,height=1,font=entry_02_label_myfont,bg='White')
entry_02_label.grid(row=1,column=1)
entry_02_label.place(x=3.5,y=165)


#button: Generate
gen_button = tk.Button(window, text='Choose a folder to run and save files', command=save_file_and_run, bg="cornflower blue", fg="White")
gen_button.grid(row=1,column=1)
gen_button.place(x=25, y=225)


#status bar
statusvar = tk.StringVar()
statusvar.set("Ready")
sbar = tk.Label(window, textvariable=statusvar, relief=tk.SUNKEN, anchor="w")
sbar.grid(row=1,column=1)
sbar.place(x=25,y=260)

# ending
window.mainloop()